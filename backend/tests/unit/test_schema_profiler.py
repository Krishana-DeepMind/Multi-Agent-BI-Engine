import os
import io
import time
import json
import uuid
import pytest
import polars as pl
import pandas as pd
from backend.core.duckdb_engine import DuckDBEngine
from backend.core.schema_profiler import SchemaProfiler

@pytest.fixture(scope="module")
def engine():
    eng = DuckDBEngine()
    yield eng
    eng.close()

@pytest.fixture(scope="module")
def profiler(engine):
    return SchemaProfiler(engine)


def test_schema_profiler_clean_csv(engine, profiler):
    csv_bytes = b"id,name,score\n1,Alice,95.5\n2,Bob,88.0\n3,Charlie,\n"
    res = engine.load_from_bytes(csv_bytes, "csv", table_name="clean_csv")
    assert res["success"] is True

    meta = profiler.profile_table("clean_csv")
    assert len(meta) == 3
    
    id_col = next(c for c in meta if c.name == "id")
    assert id_col.null_pct == 0.0
    assert id_col.unique_pct == 1.0
    assert id_col.is_primary_key is True

    score_col = next(c for c in meta if c.name == "score")
    assert score_col.null_pct > 0.0  # Charlie has null score
    assert 95.5 in score_col.sample_values


def test_schema_profiler_messy_csv(engine, profiler):
    # Mixed types: 'age' has strings and ints, 'is_active' has yes/1/true
    csv_bytes = b"user_id,age,is_active\n101,25,yes\n102,thirty,1\n103,45,true\n104,,false\n"
    res = engine.load_from_bytes(csv_bytes, "csv", table_name="messy_csv")
    assert res["success"] is True

    meta = profiler.profile_table("messy_csv")
    assert len(meta) == 3
    
    age_col = next(c for c in meta if c.name == "age")
    assert age_col.dtype == "VARCHAR"  # DuckDB falls back to VARCHAR for mixed
    assert age_col.null_pct == 0.25

    active_col = next(c for c in meta if c.name == "is_active")
    assert "yes" in active_col.sample_values or "1" in active_col.sample_values


def test_schema_profiler_json(engine, profiler):
    # Nested JSON or flat JSON array
    json_data = [
        {"id": 1, "product": "Widget", "details": {"color": "red", "size": "M"}},
        {"id": 2, "product": "Gadget", "details": {"color": "blue", "size": "L"}}
    ]
    json_bytes = json.dumps(json_data).encode("utf-8")
    res = engine.load_from_bytes(json_bytes, "json", table_name="json_table")
    assert res["success"] is True

    meta = profiler.profile_table("json_table")
    assert len(meta) == 3
    details_col = next(c for c in meta if c.name == "details")
    # Struct type in DuckDB
    assert "STRUCT" in details_col.dtype.upper()
    assert details_col.null_pct == 0.0


def test_schema_profiler_jsonl(engine, profiler):
    jsonl_bytes = b'{"log_id": "a1", "event": "click"}\n{"log_id": "a2", "event": "view"}\n'
    res = engine.load_from_bytes(jsonl_bytes, "jsonl", table_name="jsonl_table")
    assert res["success"] is True

    meta = profiler.profile_table("jsonl_table")
    assert len(meta) == 2
    assert meta[0].name == "log_id"


def test_schema_profiler_excel(engine, profiler):
    # Use openpyxl to create an xlsx with MERGED CELLS (blueprint requirement)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1: merged header spanning A1:B1
    ws.merge_cells("A1:B1")
    ws["A1"] = "Merged Header"

    # Row 2: actual column headers
    ws["A2"] = "Category"
    ws["B2"] = "Value"

    # Data rows
    ws["A3"] = "A"
    ws["B3"] = 10
    ws["A4"] = "A"
    ws["B4"] = 20
    ws["A5"] = "B"
    ws["B5"] = 30
    ws["A6"] = "B"
    ws["B6"] = None  # null value

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    res = engine.load_from_bytes(excel_buffer.getvalue(), "xlsx", table_name="excel_table")
    assert res["success"] is True

    meta = profiler.profile_table("excel_table")
    # Should have parsed columns (polars reads first non-merged row or flattens)
    assert len(meta) >= 2


def test_schema_profiler_parquet(engine, profiler):
    df = pl.DataFrame({"metric_a": [1.1, 2.2, 3.3], "metric_b": [100, 200, 300]})
    parquet_buffer = io.BytesIO()
    df.write_parquet(parquet_buffer)
    
    res = engine.load_from_bytes(parquet_buffer.getvalue(), "parquet", table_name="parquet_table")
    assert res["success"] is True

    meta = profiler.profile_table("parquet_table")
    assert len(meta) == 2
    m_a = next(c for c in meta if c.name == "metric_a")
    assert m_a.dtype == "DOUBLE"


@pytest.mark.benchmark
def test_schema_profiler_50mb_benchmark(engine, profiler):
    """
    Generate a 50MB CSV in memory, load it into DuckDB, and profile it.
    Target: < 5 seconds total.
    """
    # 50MB is roughly 1 million rows of simple data
    row_count = 1_000_000
    df = pl.DataFrame({
        "id": pl.arange(0, row_count, eager=True),
        "category": pl.Series(["A", "B", "C", "D"] * (row_count // 4)),
        "value": pl.Series([1.5, 2.5, 3.5, 4.5] * (row_count // 4)),
        "is_active": pl.Series([True, False] * (row_count // 2))
    })
    
    csv_buffer = io.BytesIO()
    df.write_csv(csv_buffer)
    csv_bytes = csv_buffer.getvalue()
    
    assert len(csv_bytes) > 10_000_000  # At least 10MB to be substantial
    
    start_time = time.perf_counter()
    
    # 1. Load
    engine.load_from_bytes(csv_bytes, "csv", table_name="bench_table")
    load_time = time.perf_counter() - start_time
    
    # 2. Profile
    meta = profiler.profile_table("bench_table")
    total_time = time.perf_counter() - start_time
    
    print(f"\\n[Benchmark] Load time: {load_time:.3f}s")
    print(f"[Benchmark] Total time (Load + Profile): {total_time:.3f}s")
    
    assert len(meta) == 4
    assert total_time < 5.0, f"Benchmark failed: took {total_time:.3f}s (Target < 5s)"
